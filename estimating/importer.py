"""Markup importer: measured Bluebeam markups -> priced takeoff line items.

Pure core (process_markups) takes injected markup dicts + factors +
job_config. ASM subjects route through the expansion engine; regular
measurement tools price as measurement x factor; unknowns -> intake.
Honors the Verified/Proposed/Rejected review gate."""
from dataclasses import dataclass, field, replace

from . import expand
from .expand import LineItem


_UNIT_BY_TYPE = {
    "Bluebeam.PDF.Annotations.AnnotationMeasurePolylength": "LF",
    "Bluebeam.PDF.Annotations.AnnotationMeasureArea": "SF",
    "Bluebeam.PDF.Annotations.AnnotationMeasureCount": "EA",
    "PolyLine": "LF", "Polygon": "SF", "Count": "EA",
}


def markup_record(raw, *, params_column="Assembly Params",
                  state_column="status"):
    """Map one Bluebeam list_markups_in_pdf entry (a dict of properties +
    custom columns) to the importer's markup-record contract. Tolerant of
    missing fields. `raw` keys are Bluebeam property names."""
    typ = raw.get("type") or raw.get("Type") or ""
    unit = _UNIT_BY_TYPE.get(typ, "?")
    meas = raw.get("measurement") or raw.get("Measurement") or 0
    try:
        meas = float(str(meas).split()[0]) if meas else 0.0
    except (ValueError, IndexError):
        meas = 0.0
    return {"subject": raw.get("subject") or raw.get("Subject") or "",
            "measurement": meas, "unit": unit,
            "params": raw.get(params_column, "") or "",
            "status": raw.get(state_column) or "Verified"}


@dataclass
class ImportResult:
    line_items: list = field(default_factory=list)
    intake: list = field(default_factory=list)
    warnings: list = field(default_factory=list)


def process_markups(markups, factors, job, *, require_verified=True):
    res = ImportResult()
    for m in markups:
        subj = m["subject"]
        status = m.get("status", "Verified")
        if status == "Rejected":
            continue
        if require_verified and status != "Verified":
            res.warnings.append(f"unverified ({status}): {subj}")
            continue
        if subj in expand._DISPATCH:
            count = int(m.get("measurement") or 1)
            try:
                items = expand.expand_marker(subj, m.get("params", ""),
                                             factors, job)
            except (KeyError, ValueError) as e:
                res.warnings.append(f"expand failed [{subj}]: {e}")
                continue
            for it in items:
                res.line_items.append(replace(
                    it, qty=round(it.qty * count, 4),
                    raw_total=round(it.raw_total * count, 2)))
        elif subj in factors:
            qty = float(m.get("measurement") or 0)
            rate = float(factors[subj])
            res.line_items.append(LineItem(subj, subj, m.get("unit", "?"),
                                           qty, rate, round(qty * rate, 2)))
        else:
            res.intake.append({"subject": subj,
                               "measurement": m.get("measurement"),
                               "unit": m.get("unit")})
    return res


def intake_rows(intake, *, line, source_date):
    """Turn unknown-subject intake records into provisional FactorRows
    (raw_cost 0.0, status provisional) for later promotion. Deduped by
    subject."""
    from .library import FactorRow
    seen, rows = set(), []
    for rec in intake:
        s = rec["subject"]
        if s in seen:
            continue
        seen.add(s)
        rows.append(FactorRow(
            subject=s, line=line, category="INTAKE", unit=rec.get("unit") or "EA",
            raw_cost=0.0, status="provisional", source="markup intake",
            source_date=source_date,
            notes="auto-captured from a takeoff markup - needs pricing"))
    return rows


def write_line_items(result, out_path, *, job, source, source_date):
    """Write the import result to an xlsx: Line Items (+ grand total),
    Intake, Warnings, and a small Job header."""
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Line Items"
    ws.append(["Component", "Subject", "Unit", "Qty", "Raw Unit $",
               "Raw Total $"])
    for c in ws[1]:
        c.font = Font(bold=True)
    total = 0.0
    for li in result.line_items:
        ws.append([li.component, li.subject, li.unit, li.qty, li.raw_unit,
                   li.raw_total])
        total += li.raw_total
    ws.append(["", "", "", "", "GRAND TOTAL (raw)", round(total, 2)])
    ws[ws.max_row][4].font = Font(bold=True)
    ws[ws.max_row][5].font = Font(bold=True)

    wi = wb.create_sheet("Intake (needs pricing)")
    wi.append(["Subject", "Measurement", "Unit"])
    for c in wi[1]:
        c.font = Font(bold=True)
    for rec in result.intake:
        wi.append([rec["subject"], rec.get("measurement"), rec.get("unit")])

    ww = wb.create_sheet("Warnings")
    ww.append(["Warning"])
    ww[1][0].font = Font(bold=True)
    for w in result.warnings:
        ww.append([w])

    wj = wb.create_sheet("Job", 0)
    for k, v in [("Source", source), ("Date", source_date),
                 ("Finish", job.get("finish_subject")),
                 ("Door", job.get("door_subject")),
                 ("Panel", job.get("panel_subject"))]:
        wj.append([k, v])
    wb.save(out_path)
