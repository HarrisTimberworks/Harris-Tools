"""HTW Factor Library xlsx: canonical Subject-keyed raw-cost rates."""
from dataclasses import dataclass, astuple

import openpyxl

FACTOR_HEADERS = ["Subject", "Line", "Category", "Unit", "Raw Cost",
                  "Status", "Source", "Source Date", "Notes"]
VENDOR_HEADERS = ["Vendor", "Preferred Source", "Discount Multiplier",
                  "Notes"]
CHANGELOG_HEADERS = ["Date", "Version", "Author", "Change"]
VALID_STATUS = {"active", "provisional", "retired"}
VALID_LINE = {"R", "C", "Both"}


@dataclass
class FactorRow:
    subject: str
    line: str
    category: str
    unit: str
    raw_cost: float
    status: str
    source: str
    source_date: str
    notes: str


def create_library(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Factors"
    ws.append(FACTOR_HEADERS)
    wb.create_sheet("Vendors").append(VENDOR_HEADERS)
    wb.create_sheet("Changelog").append(CHANGELOG_HEADERS)
    wb.save(path)


def validate(rows):
    seen = set()
    for r in rows:
        if r.subject in seen:
            raise ValueError(f"duplicate subject: {r.subject!r}")
        seen.add(r.subject)
        if r.status not in VALID_STATUS:
            raise ValueError(f"bad status {r.status!r} on {r.subject!r}")
        if r.line not in VALID_LINE:
            raise ValueError(f"bad line {r.line!r} on {r.subject!r}")
        if not isinstance(r.raw_cost, (int, float)) or isinstance(
                r.raw_cost, bool) or r.raw_cost < 0:
            raise ValueError(
                f"bad raw_cost {r.raw_cost!r} on {r.subject!r} — "
                f"must be a number >= 0")


def write_factors(path, rows):
    validate(rows)
    wb = openpyxl.load_workbook(path)
    ws = wb["Factors"]
    ws.delete_rows(2, ws.max_row)
    for r in rows:
        ws.append(astuple(r))
    wb.save(path)


def load_factors(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Factors"]
    rows = []
    for vals in ws.iter_rows(min_row=2, values_only=True):
        padded = (tuple(vals) + (None,) * 9)[:9]
        if padded[0] is None:
            continue
        cells = [v if v is not None else "" for v in padded]
        try:
            cells[4] = float(cells[4])
        except (TypeError, ValueError):
            raise ValueError(
                f"Raw Cost for subject {cells[0]!r} is not a number: "
                f"{cells[4]!r}")
        rows.append(FactorRow(*cells))
    wb.close()
    return rows


def append_changelog(path, *, version, author, change, date):
    wb = openpyxl.load_workbook(path)
    wb["Changelog"].append([date, version, author, change])
    wb.save(path)


def latest_version(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Changelog"]
    version = None
    for vals in ws.iter_rows(min_row=2, values_only=True):
        if vals[1] is not None:
            version = vals[1]
    wb.close()
    return version
