import pytest
from estimating import importer

FACTORS = {
    "CASE-COMM - Base PLam - LF": 210.0,
    "FF FinEnds - Flush": 35.0,
    "FIN - Stain (1 Sided)": 2.0,
    "DOOR - Slab - Paint Grade": 18.0,
    "Panels - Paint Grade": 9.0,
}
JOB = {"finish_subject": "FIN - Stain (1 Sided)",
       "door_subject": "DOOR - Slab - Paint Grade",
       "panel_subject": "Panels - Paint Grade"}


def _m(subject, measurement, unit, params="", status="Verified"):
    return {"subject": subject, "measurement": measurement, "unit": unit,
            "params": params, "status": status}


def test_regular_measurement_priced_directly():
    r = importer.process_markups(
        [_m("CASE-COMM - Base PLam - LF", 12.5, "LF")], FACTORS, JOB)
    assert len(r.line_items) == 1
    li = r.line_items[0]
    assert li.subject == "CASE-COMM - Base PLam - LF"
    assert li.qty == 12.5
    assert li.raw_total == pytest.approx(2625.0)
    assert r.intake == [] and r.warnings == []


def test_asm_marker_expands():
    r = importer.process_markups(
        [_m("ASM - Finished End (FF Flush) - EA", 1, "EA", "D=24 H=84")],
        FACTORS, JOB)
    subs = {li.subject for li in r.line_items}
    assert "FF FinEnds - Flush" in subs
    assert "FIN - Stain (1 Sided)" in subs


def test_asm_count_multiplies_expansion():
    one = importer.process_markups(
        [_m("ASM - Finished End (FF Flush) - EA", 1, "EA", "D=24 H=84")],
        FACTORS, JOB)
    three = importer.process_markups(
        [_m("ASM - Finished End (FF Flush) - EA", 3, "EA", "D=24 H=84")],
        FACTORS, JOB)
    assert sum(li.raw_total for li in three.line_items) == pytest.approx(
        sum(li.raw_total for li in one.line_items) * 3)


def test_unknown_subject_goes_to_intake():
    r = importer.process_markups(
        [_m("CUSTOM - Mystery Thing - EA", 2, "EA")], FACTORS, JOB)
    assert r.line_items == []
    assert r.intake == [{"subject": "CUSTOM - Mystery Thing - EA",
                         "measurement": 2, "unit": "EA"}]


def test_rejected_skipped_proposed_warned():
    r = importer.process_markups([
        _m("CASE-COMM - Base PLam - LF", 5, "LF", status="Rejected"),
        _m("CASE-COMM - Base PLam - LF", 7, "LF", status="Proposed"),
    ], FACTORS, JOB)
    assert r.line_items == []
    assert any("Proposed" in w for w in r.warnings)
    assert not any("Rejected" in w for w in r.warnings)


def test_expand_failure_becomes_warning_not_crash():
    r = importer.process_markups(
        [_m("ASM - Open Interior - EA", 1, "EA", "W=36 H=84 D=24 SH=3")],
        {}, JOB)
    assert r.line_items == []
    assert any("Open Interior" in w for w in r.warnings)


def test_intake_rows_are_provisional_factorrows():
    from estimating import library
    intake = [{"subject": "CUSTOM - New Thing - EA", "measurement": 4,
               "unit": "EA"}]
    rows = importer.intake_rows(intake, line="C", source_date="2026-06-12")
    assert len(rows) == 1
    r = rows[0]
    assert isinstance(r, library.FactorRow)
    assert r.subject == "CUSTOM - New Thing - EA"
    assert r.line == "C"
    assert r.status == "provisional"
    assert r.raw_cost == 0.0
    assert r.unit == "EA"
    assert "intake" in r.source.lower()


def test_intake_dedupes_repeated_subject():
    intake = [{"subject": "X - A - EA", "measurement": 1, "unit": "EA"},
              {"subject": "X - A - EA", "measurement": 2, "unit": "EA"}]
    rows = importer.intake_rows(intake, line="C", source_date="2026-06-12")
    assert len(rows) == 1


def test_write_line_items_xlsx(tmp_path):
    import openpyxl
    r = importer.process_markups(
        [_m("CASE-COMM - Base PLam - LF", 12.5, "LF")], FACTORS, JOB)
    out = tmp_path / "takeoff_lines.xlsx"
    importer.write_line_items(r, out, job=JOB,
                              source="Test Job", source_date="2026-06-12")
    wb = openpyxl.load_workbook(out)
    assert "Line Items" in wb.sheetnames
    ws = wb["Line Items"]
    headers = [c.value for c in ws[1]]
    assert headers == ["Component", "Subject", "Unit", "Qty",
                       "Raw Unit $", "Raw Total $"]
    vals = [r2[5] for r2 in ws.iter_rows(min_row=2, values_only=True)
            if r2[5] is not None]
    assert 2625.0 in vals


def test_write_includes_intake_and_warnings_sheets(tmp_path):
    import openpyxl
    r = importer.process_markups([
        _m("CUSTOM - Mystery - EA", 2, "EA"),
        _m("CASE-COMM - Base PLam - LF", 1, "LF", status="Proposed"),
    ], FACTORS, JOB)
    out = tmp_path / "t.xlsx"
    importer.write_line_items(r, out, job=JOB, source="J", source_date="d")
    wb = openpyxl.load_workbook(out)
    assert "Intake (needs pricing)" in wb.sheetnames
    assert "Warnings" in wb.sheetnames
